from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import sys
import openpyxl
import time
import services

#função que cria a string com o texto para a pessoa // recebe o nome da pessoa e o servico que deseja comprar
def texto(name, service):
    return '''Bom dia ''' + name + ''', tudo bem?

Me chamo Pedro e hoje em dia trabalho na Empresa Júnior PUC-Rio, uma consultoria formada por jovens, estudantes da própria universidade. Realizamos diversos serviços - nas áreas de Marketing, Audiovisual, Arquitetura, Processos, Finanças, Assessoria de Imprensa e Design - com o intuito de alavancar empresas e fomentar o desenvolvimento de empreendimentos.

Me identifico muito com a marca e com todo o posicionamento que vocês prezam e propagam no mercado atualmente. Por isso, gostaria de ajudar a impulsionar mais o seu desenvolvimento e expansão. Acredito que  poderíamos dialogar o diferencial da Empresa Júnior com a singularidade de vocês.

Você é a pessoa ideal para conversarmos sobre o assunto?
'''

#função que escreve o email
def escrever_email(index):
    #usa a parte do assunto como gatilho para escrever o email
    asst = driver.find_element_by_class_name("aoT")
    asst.send_keys(Keys.TAB + texto(cont_na[index], cont_se[index]) + Keys.DELETE * 4 + Keys.ENTER)

    #ativa o marcador de prospecção
    driver.find_element_by_id(':n0').click()
    driver.find_element_by_id(':mx').click()
    driver.find_element_by_id(':ta').click()
    driver.find_element_by_id(':n0').click()
    driver.find_element_by_id(':op').click()

#função que abre nova janela com a url do gmail mailto
def preparar_janela(a):
    driver.execute_script("window.open('');")
    driver.switch_to.window(driver.window_handles[t])
    driver.get('https://mail.google.com/mail/u/0/?view=cm&fs=1&tf=1&source=mailto&su=Contato+|+Empresa+Júnior+PUC-Rio&to=' + cont_em[a])
    time.sleep(3)

#cria lista com o email de cada contato
cont_em = []

#cria lista com o nome de cada contato
cont_na = []

#cria lista com o servico de cada contato
cont_se = []

#cria array que retorna o nome do serviço baseado no seu id
servs = services.serv_retrn()

#abre o excel e pega os contatos que estão lá
workbook = openpyxl.load_workbook('/Users/pedroenriqueandrade/Desktop/pymailer/Mailer/db.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')

print('\nOs contatos analisados na planilha foram:')
qtd_c = 0
i = 4
#mostra no terminal os emails que serao criados ate a primeira celula estar vazia e pede a confirmação no final
while sheet.cell(row=i, column=3).value:
    cont_na.append((sheet.cell(row=i, column=3)).value)
    cont_em.append((sheet.cell(row=i, column=5)).value)
    cont_se.append((sheet.cell(row=i, column=6)).value)
    print(cont_na[qtd_c] + ' | ' + cont_em[qtd_c] + ' | ' + servs[cont_se[qtd_c]])
    qtd_c = qtd_c + 1
    i = i + 1

print('\nDeseja confirmar esses contatos acima?')
resp1 = input('Y ou N\n')

while (resp1 != 'Y') and (resp1 != 'N'):
    print('\npor favor responda apenas com Y ou N')
    print('Deseja confirmar esses contatos acima?')
    resp1 = input('Y ou N\n')

if resp1 != 'Y':
    sys.exit('Paralizando as operacoes...\n')


driver = webdriver.Chrome('/Users/pedroenriqueandrade/Desktop/pymailer/drivers/chromedriver')
time.sleep(1)
driver.get("https://accounts.google.com/AccountChooser?service=mail&continue=https://mail.google.com/mail/")

#colocar o email
login = driver.find_element_by_id('identifierId')
login.send_keys('pandrade@empresajunior.com.br')

#dar o ok após colocar o email
login_n = driver.find_element_by_id('identifierNext')
login_n.click()

time.sleep(1.8)

#colocar a senha
psw = driver.find_element_by_name('password')
psw.send_keys('pefa1997')

#dar o ok após colocar a senha
psw_n = driver.find_element_by_id('passwordNext')
psw_n.click()

time.sleep(1)
#espera o gmail entrar e carregar a pagina do email
driver.get('https://mail.google.com/mail/u/0/?view=cm&fs=1&tf=1&source=mailto&su=Contato+|+Empresa+Júnior+PUC-Rio&to=' + cont_em[0])
time.sleep(3)
escrever_email(0)

#caso o numero de contatos (armazenado em qtd_c) seja maior que 1, todo o processo deve ser feito de novo
if (qtd_c > 1):
    for t in range (1, qtd_c):
        preparar_janela(t)
        escrever_email(t)
