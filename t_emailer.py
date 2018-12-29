import openpyxl
import pyperclip
import sys
import webbrowser


print('\n\nBem vindo ao mailer\n\nCriado por Pedro Andrade durante as suas ferias no final de 2018\n')
print('Tenha nocao de que todos os contatos deverao estar listados na planilha db.xlsx')
print('#######################')
print('Iniciando o codigo')
print('#######################')
print('Quantos contatos serao analisados?')

qtd_c = input()
qtd_c = int(qtd_c)

#caso a pessoa não inputar um numero inteiro positivo
if qtd_c <= 0:
    sys.exit('A quantidade de contatos deve ser de no minimo 1\n\nParalizando as operacoes...\n')


workbook = openpyxl.load_workbook('/Users/pedroenriqueandrade/Desktop/pymailer/db.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')

for i in range(4,(4+qtd_c)):
    webbrowser.open('https://mail.google.com/mail/u/0/?view=cm&fs=1&tf=1&source=mailto&su=Contato+|+Empresa+Júnior+PUC-Rio&to=' + (sheet.cell(row=i, column=5)).value)
    print('Copiamos o corpo do email para o seu clipboard')
    pyperclip.copy('''Bom dia ''' + (sheet.cell(row=i, column=3)).value + ''', tudo bem?

Me chamo Pedro e hoje em dia trabalho na Empresa Júnior PUC-Rio, uma consultoria formada por jovens, estudantes da própria universidade. Realizamos diversos serviços - nas áreas de Marketing, Audiovisual, Arquitetura, Processos, Finanças, Assessoria de Imprensa e Design - com o intuito de alavancar empresas e fomentar o desenvolvimento de empreendimentos.

Me identifico muito com a marca e com todo o posicionamento que vocês prezam e propagam no mercado atualmente. Por isso, gostaria de ajudar a impulsionar mais o seu desenvolvimento e expansão. Acredito que  poderíamos dialogar o diferencial da Empresa Júnior com a singularidade de vocês.

Você é a pessoa ideal para conversarmos sobre o assunto? \n
''')
    spam = pyperclip.paste()
    print('Va ate a pagina que acabou de ser aberta e cole o texto la')
    print('Assim que você estiver pronto aperte ENTER')
    input()
