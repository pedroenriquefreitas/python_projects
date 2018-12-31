from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import sys
import openpyxl
import time


#cria lista com o email de cada contato
cont_em = []

#cria lista com o nome de cada contato
cont_na = []

#abre o excel e pega os contatos que estão lá
workbook = openpyxl.load_workbook('/Users/pedroenriqueandrade/Desktop/pymailer/db.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')

#é necessario perguntar antes quantos emails serao enviados
qtd_c = input('Quantos contatos foram colocados na planilha?\n')
qtd_c = int(qtd_c)

#caso a pessoa não inputar um numero inteiro positivo
if qtd_c <= 0:
    sys.exit('A quantidade de contatos deve ser de no minimo 1\n\nParalizando as operacoes...\n')

#mostra no terminal os emails que serao criados e pede para ser confirmado
print('\nOs contatos analisados na planilha foram:')
for i in range(4,(4+qtd_c)):
    cont_na.append((sheet.cell(row=i, column=3)).value)
    cont_em.append((sheet.cell(row=i, column=5)).value)
    print(cont_na[i] + ' | ' + cont_em[i])

print('\nDeseja confirmar esses contatos acima?')
resp1 = input('Y ou N\n')

while (resp1 != 'Y') and (resp1 != 'N'):
    print('\npor favor responda apenas com Y ou N')
    print('Deseja confirmar esses contatos acima?')
    resp1 = input('Y ou N\n')

if resp1 != 'Y':
    sys.exit('Paralizando as operacoes...\n')


driver = webdriver.Chrome('/Users/pedroenriqueandrade/Desktop/pymailer/drivers/chromedriver')
time.sleep(1)
driver.get('https://mail.google.com/mail/u/0/?view=cm&fs=1&tf=1&source=mailto&su=Contato+|+Empresa+Júnior+PUC-Rio&to=' + cont_em[0])

#colocar o email
login = driver.find_element_by_id('identifierId')
login.send_keys('pandrade@empresajunior.com.br')

#dar o ok após colocar o email
login_n = driver.find_element_by_id('identifierNext')
login_n.click()

time.sleep(1.8)

#colocar a senha
psw = driver.find_element_by_name('password')
psw.send_keys('pefa1997')

#dar o ok após colocar a senha
psw_n = driver.find_element_by_id('passwordNext')
psw_n.click()

#espera o gmail entrar e carregar a pagina do email
time.sleep(6.5)

#coloca o assunto
asst = driver.find_element_by_class_name("aoT")
#asst.send_keys('Contato | Empresa Júnior PUC-Rio')
asst.send_keys(Keys.TAB + '''Bom dia ''' + cont_na[0] + ''', tudo bem?

Me chamo Pedro e hoje em dia trabalho na Empresa Júnior PUC-Rio, uma consultoria formada por jovens, estudantes da própria universidade. Realizamos diversos serviços - nas áreas de Marketing, Audiovisual, Arquitetura, Processos, Finanças, Assessoria de Imprensa e Design - com o intuito de alavancar empresas e fomentar o desenvolvimento de empreendimentos.

Me identifico muito com a marca e com todo o posicionamento que vocês prezam e propagam no mercado atualmente. Por isso, gostaria de ajudar a impulsionar mais o seu desenvolvimento e expansão. Acredito que  poderíamos dialogar o diferencial da Empresa Júnior com a singularidade de vocês.

Você é a pessoa ideal para conversarmos sobre o assunto?
''' + Keys.DELETE * 4 + Keys.ENTER)

#ativa o marcador de prospecção
driver.find_element_by_id(':n0').click()
driver.find_element_by_id(':mx').click()
driver.find_element_by_id(':ta').click()

if (qtd_c > 1):
    for t in range (1, qtd_c):
        scrpt = '''window.open("https://mail.google.com/mail/u/0/?view=cm&fs=1&tf=1&source=mailto&su=Contato+|+Empresa+Júnior+PUC-Rio&to="''' + cont_em[t] + ''',"_blank");'''
        driver.execute_script(scrpt)
